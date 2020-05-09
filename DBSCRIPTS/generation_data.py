from openpyxl import load_workbook
import re
import pandas as pd
from itertools import groupby
import timeit


class generation_data(object):
    def __init__(self, path_from, path_to, rb):
        self.path_from = path_from
        self.path_to = path_to
        self.rb = rb

    def cvs_create(self,table):
        df = pd.DataFrame(table)
        df.to_csv(self.path_to,index=False)

    def xlsx_create(self, dfun, sheet_names, n):
        if n == 1:
            df = dfun
            df.to_excel(self.path_to, sheet_name=sheet_names)
        else:
            df1 = dfun[0]
            df2 = dfun[1]
            with pd.ExcelWriter(self.path_to) as writer:
                df1.to_excel(writer, sheet_name=sheet_names[0])
                df2.to_excel(writer, sheet_name=sheet_names[1])

    def read_data(self, type):
        path = self.path_from
        if type == 'xlsx':
            wb = load_workbook(path)
            sheetname = wb.sheetnames
            sheet = wb[sheetname[0]]
            result = sheet
        else:
            result = pd.read_csv(path, sep=',')#или ;
        return result

    def generation_csv_table(self):
        path=self.path_from
        result = pd.read_csv(path, sep=',', encoding='utf-8').to_dict()
        #print(result)
        #TITLE = list(result.keys())
        #print(TITLE)
        
        keys = ['ID_SP_NAR',
        'PersID',
        'машинист инструктор',
        'RoadID',
        'SpsSerieID',
        'SpsGroupID',
        'EnterpriseID',
        'DateFrom',
        'TabNum',
        'fired',
        'LastName',
        'FirstName',
        'PatrName',
        'KOD_DEPO',
        'CurrTabNum'
        ]


        list1 = list(result.get('ID_SP_NAR').values())
        list2 = list(result.get('PersID').values())
        list3 = list(result.get('машинист инструктор').values())
        list4 = list(result.get('RoadID').values())
        list5a=list(result.get('SpsSerieID').values())
        list5b = list(result.get('SpsGroupID').values())
        list5 = [str(list5a[i])+str(list5b[i]) for i in range(len(list5a))]
        list7 = list(result.get('EnterpriseID').values())
        list8 = list(result.get('DateFrom').values())
        
        list9 = list(result.get('TabNum').values())
        list10 = list(result.get('fired').values())
        
        list11 = list(result.get('LastName').values())
        list12 = list(result.get('FirstName').values())
        list13 = list(result.get('PatrName').values())
        list14 = list(result.get('KOD_DEPO').values())
        list15 = list(result.get('CurrTabNum').values())

        pers_id = [list1, list2, list3, list4, list5a, list5b, list7, list8, list9, list10, list11,
                   list12, list13, list14, list15]

#        pers_id = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12]
        PERS_ID = {keys[i]: pers_id[i] for i in range(len(keys))}

        return PERS_ID

    def generation_general_table(self, sheet):
        TITLE = []
        i = 1
        while sheet.cell(row=1, column=i).value != None:  # title
            TITLE.append(sheet.cell(row=1, column=i).value)
            i += 1

        need_id_pers = ([i for i in range(len(TITLE)) if TITLE[i] == 'PersID'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'машинист инструктор'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'RoadID'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'KokrsID'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'EnterpriseID'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'LastName'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'FirstName'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'PatrName'],
                        # [i for i in range(len(TITLE)) if TITLE[i]=='SpsID'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'KOD_DEPO'],
                        [i for i in range(len(TITLE)) if TITLE[i] == 'TabNum'])
        # print(need_id_pers)
        id_need_id_status = [i for i in range(len(TITLE)) if TITLE[i] == 'fired']

        id_ID_SP_NAR = [i for i in range(len(TITLE)) if TITLE[i] == 'ID_SP_NAR']
        id_Date_NAR = [i for i in range(len(TITLE)) if TITLE[i] == 'Date_NAR']

        pers_id = []
        i = 2
        PERS_ID = {}
        keys = ['PersID',
                'машинист инструктор',
                'RoadID',
                'KokrsID',
                'EnterpriseID',
                'LastName',
                'FirstName',
                'PatrName',
                'KOD_DEPO',
                'TabNum']

        while sheet.cell(row=i, column=need_id_pers[0][0]).value != None:  # title
            pers_id.append([sheet.cell(row=i, column=need_id_pers[0][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[1][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[2][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[3][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[4][1] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[5][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[6][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[7][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[8][0] + 1).value,
                            sheet.cell(row=i, column=need_id_pers[9][0] + 1).value])

            i += 1
        l = len(pers_id)

        PERS_ID = {keys[i]: [pers_id[j][i] for j in range(len(pers_id))] for i in range(len(keys))}

        ID_SP_NAR = []
        i = 2
        while sheet.cell(row=i, column=id_ID_SP_NAR[0]).value != None:  # title
            ID_SP_NAR.append(sheet.cell(row=i, column=id_ID_SP_NAR[0] + 1).value)
            i += 1

        idspnar = {'ID_SP_NAR': ID_SP_NAR}
        PERS_ID.update(idspnar)

        id_status = []
        i = 2
        while sheet.cell(row=i, column=id_need_id_status[0]).value != None:  # 1 уволен
            id_status.append(sheet.cell(row=i, column=id_need_id_status[0] + 1).value)
            i += 1

        status = {'fired': id_status}
        PERS_ID.update(status)

        # Date_NAR = []
        # i = 2
        # while sheet.cell(row=i, column=id_Date_NAR[0]).value != None:  # title
        #   Date_NAR.append(sheet.cell(row=i, column=id_Date_NAR[0] + 1).value)
        #   i += 1

        # datenar={'DATE_NAR':Date_NAR}
        # PERS_ID.update(datenar)

        return PERS_ID

    def table_sorted(self, keys, table):
        PERS_ID = {keys[i]: table[i] for i in range(len(keys))}
        return PERS_ID

    def group_ID(self, l):
        return groupby(sorted(l))

    def tab_sort(self, list_d):
        list_l = list_d.copy()
        l = []
        pp = []
        id_i = 0
        for index, el_1 in enumerate(list_l):#DEBUG
            #print('element sort', index, 'of', len(list_l)) 
            if all(el_0 != el_1 for el_0 in l):
                l.append(el_1)
                k = 0
                kj = id_i + 1
                p = [id_i]
                for el_2 in list_l[id_i + 1:len(list_l)]:
                    if el_2 == el_1:
                        p.append(kj)
                        l.append(el_2)
                    kj += 1
                    k += 1
                pp.append(p)
            id_i += 1
        keys = [list(groupby(l))[i][0] for i in range(len(list(groupby(l))))]
        dict = {str(keys[i]): pp[i] for i in range(len(keys))}
        return [dict, keys, pp, l]

    def sorted_pair(self, list2, list1):
#        list1_new = self.tab_sort(list1)[3]
        ll = self.tab_sort(list1)
        numer = [ll[2][i][j] for i in range(len(ll[2])) for j in range(len(ll[2][i]))]
        list3 = [list2[numer[i]] for i in range(len(list2))]
        return [list3, list1]

    def sorted_personal(self, table):
        tab = table.get('TabNum')
        # tab=self.tab_sort(tab)[0]
        keys = table.keys()
        VAR = []
        for k in keys:
            var = table.get(k)
            print("TAB", str(k))
            var = self.sorted_pair(var, tab)[0] #DEBUG
            VAR.append(var)
        return VAR

    def sorted_personal_pers_id(self, table):
        tab = table.get('PersID')
        # tab=self.tab_sort(tab)[0]
        keys = table.keys()
        VAR = []
        for k in keys:
            var = table.get(k)
            var = self.sorted_pair(var, tab)[0]
            VAR.append(var)
        return VAR

    def probability_calculation(self, list_nar, fired):
        if len(list_nar) > 0:
            #l_nar = self.tab_sort(list_nar)[3]

            probability = [round(100 * list_nar.count(ln) / len(list_nar), 1) for ln in list_nar]
#            probability = [probability[i] * (1 - fired[i]) for i in range(len(probability))]
        else:
            probability = [0 for ln in list_nar]
        return probability

    def table_probability(self, table):
        list_nar = table[0] #10
        fired = table[9] #11
        tab = table[8] #9
        # print(fired)
        i = 0
        probability = []
        for t in groupby(tab):
            lnar = []
            f = []
            for el in t[1]:
                # print(el)
                lnar.append(list_nar[i])
                f.append(fired[i])
                i += 1
            pr = self.probability_calculation(lnar, f)
            probability.append(pr)
        #print(probability)
        prob = [probability[i][j] for i in range(len(probability)) for j in range(len(probability[i]))]
        return prob

    def table_probability_pers_id(self, table):
        list_nar = table[10]
        fired = table[11]
        tab = table[0]
        # print(fired)
        i = 0
        probability = []
        for t in groupby(tab):
            lnar = []
            f = []
            for el in t[1]:
                # print(el)
                lnar.append(list_nar[i])
                f.append(fired[i])
                i += 1
            pr = self.probability_calculation(lnar, f)
            probability.append(pr)
        print(probability)
        prob = [probability[i][j] for i in range(len(probability)) for j in range(len(probability[i]))]
        return prob

    def probability_general(self, table):
        list_nar = table[0] # 10
        i = 0
        prob = []
        for lnar in list_nar:
            prob.append(round(100 * list_nar.count(lnar) / len(list_nar), 1))
            i += 1
        return prob

    def update_table(self, table, dict):
        table.update(dict)
        return table

    def DEF_NAR(self, idnar):
        df = pd.read_csv(self.rb, sep='\t', error_bad_lines=False, encoding='utf-8', decimal=',')  # раасшифровка
        lest = list(df["ID_SP_NAR;NAME"][:269])
        l_split = [re.split(r';', lest[i]) for i in range(len(lest))]
        l_num = [int(l_split[i][0]) for i in range(len(l_split) - 2)]
        l_text = [l_split[i][1] for i in range(len(l_split) - 2)]
        resp = 'NONE'
        for i in range(len(l_num)):
            if int(l_num[i]) == int(idnar): #1128 # int(idnar): #DEBUG
                resp = l_text[i]
        return resp

    def tab_nar(self, l_nar):
        return [self.DEF_NAR(ln) for ln in l_nar]

    def sort_list_by(self,list1, list2):
        zipped_pairs = zip(list2, list1)
        z = [x for x, _ in groupby(sorted(zipped_pairs))]
        return z

    def group_sorted_list(self,x, y):
        l = list(self.sort_list_by(x, y))
        # print(l)
        ll = []
        ll.append(l[0][1])
        for i in range(1, len(l)):
            if l[i][0] != l[i - 1][0]:
                ll.append(l[i][1])
        return ll


if __name__ == "__main__":
    start = timeit.default_timer()
    lasttime = start

    PATH = 'sample/sample_3.xlsx'
    path_output_data = 'results_update.xlsx'
    path_rb = 'data/rb.csv'
    keys = ['PersID',
            'машинист инструктор',
            'RoadID',
            'KokrsID',
            'EnterpriseID',
            'LastName',
            'FirstName',
            'PatrName',
            'KOD_DEPO',
            'TabNum',
            'ID_SP_NAR',
            'fired',
            'personal probability',
            'general probability',
            'расшифровка']
    data = generation_data(path_from=PATH, path_to=path_output_data, rb=path_rb)
    sheet = data.read_data('xlsx')
    d = data.generation_general_table(sheet)

    d_sort = data.sorted_personal(d)  # сортировка по табельнику
    probability = data.table_probability_pers_id(d_sort)
    d_sort.append(probability)

    probability_general = data.probability_general(d_sort)
    d_sort.append(probability_general)
    crypt = data.tab_nar(d_sort[10])
    d_sort.append(crypt)
    d_sort = data.table_sorted(keys, d_sort)
    sheet_name_tab = 'сортировка по табельнику'
    # data.xlsx_create(d_sort,sheet_name_tab)

    df = pd.DataFrame(d_sort)
    data.xlsx_create(df, sheet_name_tab, 1)

    stop = timeit.default_timer()
    print('Time: ', stop - start)

    '''d_sort1 = data.sorted_personal_pers_id(d)  # сортировка по ID
    probability1 = data.table_probability_pers_id(d_sort1)
    d_sort1.append(probability1)
    crypt1 = data.tab_nar(d_sort1[0])
    d_sort1.append(crypt1)
    d_sort1 = data.table_sorted(keys, d_sort1)
    sheet_name_tab1 = 'сортировка по ID'

    df = [pd.DataFrame(d_sort),pd.DataFrame(d_sort1)]
    sheet_names=[sheet_name_tab,sheet_name_tab1]
    data.xlsx_create(df, sheet_names)'''
