from math import *
from tkinter import *#this library is shit
import openpyxl
import pandas as pd
from itertools import groupby

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)

from tkinter.ttk import Style
from PIL import Image, ImageTk


PATH='update_0.csv'

def read_data_excel(path, column=14):  # column=14,15
    wb = openpyxl.load_workbook(path)
    sheetname = wb.sheetnames
    sheet = wb[sheetname[0]]
    rows = []
    i = 2
    while sheet.cell(row=i, column=column).value != None:  # title
        rows.append([sheet.cell(row=i, column=column).value])
        i += 1
    return rows

def read_data(path,name):
    res=pd.read_csv(path, sep=',').to_dict()
    list_col=list(res.get(name).values())
    return list_col


def sort_list(list1, list2):
    zipped_pairs = zip(list2, list1)
    z = [x for _, x in sorted(zipped_pairs)]
    return z


def sort_list_by(list1, list2):
    zipped_pairs = zip(list2, list1)
    z = [x for x, _ in groupby(sorted(zipped_pairs))]
    return z

def group_sorted_list(x,y):
    l = list(sort_list_by(x, y))
    #print(l)
    ll = []
    ll.append(l[0][1])
    for i in range(1,len(l)):
        if l[i][0] != l[i-1][0]:
            ll.append(l[i][1])
    return ll

def list_tab_name(path):
    #tab = read_data(path, column=11)  # табельник
    tab = read_data(path, 'TabNum')  # табельник
    len_nar = len(tab)
    tab = [tab[i] for i in range(len_nar)]
    tab1=tab.copy()
    tab2=tab.copy()
    tab3=tab.copy()
    tab4=tab.copy()

    FirstName = read_data(path, 'FirstName')  # FirstName
    #FirstName = [str(FirstName[i]) for i in range(len_nar)]
    FirstName=group_sorted_list(FirstName,tab1)

    LastName = read_data(path, 'LastName')  # LastName
    #LastName = [str(LastName[i]) for i in range(len_nar)]
    LastName = group_sorted_list(LastName,tab2)

    PatrName = read_data(path, 'PatrName')  # PatrName
    #PatrName = [str(PatrName[i]) for i in range(len_nar)]
    PatrName = group_sorted_list(PatrName,tab3)

    PersID = read_data(path, 'PersID')  # PersID
    #PersID = [PersID[i] for i in range(len_nar)]
    PersID = group_sorted_list(PersID,tab4)

    tab5=[list(groupby(sorted(tab)))[i][0] for i in range(len(list(groupby(sorted(tab)))))]

    print(FirstName)

    return [tab5,FirstName,LastName,PatrName,PersID]

def calc_probability(tabnum='False'):
    path = PATH
    print('start calculation')
    print(tabnum)
    if tabnum == 'False':
        pr = read_data(path, 'general probability')
        len_nar = len(pr)
        #probabilty=pr
        probabilty = [float(pr[i]) for i in range(len_nar)]
        #print(probabilty)

        tab = read_data(path, 'TabNum')  # табельник
        #tab = [tab[i] for i in range(len_nar)]

        ID_SP_NAR = read_data(path, 'ID_SP_NAR')  # нарушения
        #ID_SP_NAR = [ID_SP_NAR[i] for i in range(len_nar)]

        crypt = read_data(path, 'расшифровка')  # расшифровка
        #crypt = [crypt[i] for i in range(len_nar)]

        idspnar = ID_SP_NAR.copy()
        idspnar2 = ID_SP_NAR.copy()
        idspnar1 = ID_SP_NAR.copy()
        pr = probabilty.copy()

        pr = group_sorted_list(pr, idspnar)
        print(pr)

        crypt = group_sorted_list(crypt, idspnar2)
        #print(crypt)

        tab = group_sorted_list(tab, idspnar1)
        #print(tab)

        l=list(groupby(sorted(ID_SP_NAR)))
        ID_SP_NAR=[l[i][0] for i in range(len(l))]
        #print(ID_SP_NAR)
        print('end calculation 1')
        diagramma_result(pr, crypt, tabnum)
        return [pr, ID_SP_NAR,crypt,tab]
    else:
        pr = read_data(path, 'personal probability')
        len_nar=len(pr)
        #probabilty=pr
        probabilty = [pr[i] for i in range(len_nar)]

        tab = read_data(path, 'TabNum')  # табельник
        #tab = [tab[i] for i in range(len_nar)]

        ID_SP_NAR = read_data(path, 'ID_SP_NAR')  # нарушения
        #ID_SP_NAR = [ID_SP_NAR[i] for i in range(len_nar)]

        crypt = read_data(path, 'расшифровка')  # расшифровка
        #crypt = [crypt[i] for i in range(len_nar)]

        pr_tabnum = []
        pr_ID_SP_NAR = []
        pr_crypt = []
        i = 0
        #print(tabnum)
        #if all(int(tab[i])!=tabnum for i in range(len(tab))):
         #   pass

        for t in tab:
            if int(t) == int(tabnum):

                pr_ID_SP_NAR.append(ID_SP_NAR[i])
                pr_crypt.append(crypt[i])
                pr_tabnum.append(probabilty[i])
            i += 1
        idspnar = pr_ID_SP_NAR.copy()
        idspnar1=pr_ID_SP_NAR.copy()
        idspnar2 = pr_ID_SP_NAR.copy()
        print(pr_tabnum)

        pr_tabnum = group_sorted_list(pr_tabnum, idspnar)
        print(pr_tabnum)
        pr_crypt = group_sorted_list(pr_crypt, idspnar1)
        pr_ID_SP_NAR=group_sorted_list(pr_ID_SP_NAR,idspnar2)
        print('...end calculation 2.')
        diagramma_result(pr_tabnum,pr_crypt,tabnum)
        return [pr_tabnum, pr_crypt,pr_ID_SP_NAR]

def diagramma_result(probability,group,name):#https://eax.me/python-matplotlib/
    print('draw plot...')
    data_names = group
    data_values = probability

    dpi = 45
    fig = plt.figure(dpi=dpi, figsize=(150 / dpi, 150 / dpi))
    mpl.rcParams.update({'font.size': 8})

    plt.title(name)

    xs = range(len(data_names))

    plt.pie(
        data_values, autopct='%.1f', radius=1.1,
        explode=[0.05] + [0 for _ in range(len(data_names) - 1)])
    plt.legend(
        bbox_to_anchor=(-0.15, 0.05, 0.05, 0.05),
        loc='lower left', labels=data_names)

    fig.savefig('диаграмма.png', dpi=100)
    print('... drawing is finishing.')

    style = Style()
    style.configure("TFrame", background="#333")

    im = Image.open("диаграмма.png")
    img = ImageTk.PhotoImage(im)
    label3 = Label(image=img)
    label3.image = img
    label3.place(x=500, y=100)

root = Tk()
root.title("статистика нарушения машинистов")
root.geometry("1200x600")
root.config(background="lightblue")

# frame=Frame(root, width=1200, height=600, background="lightblue")
#c = Canvas(root,width = 1200, height = 1200, bg = "lightblue", cursor = "pencil")
#L1 = Label(width=50, text="наиболее вероятные нарушения машинистов").grid(row=0, column=0)

t=Label(root,width=50, text="наиболее вероятные нарушения машинистов")
t.config(font=("Times", 12),background="lightblue")
#t.pack(side=BOTTOM,expand=1)
t.place(x=400, y=20)


OptionList = [
    "по всей дирекции",
    "По машинистам",
    "По дороге"]

variable = StringVar(root)
variable.set(OptionList[0])

w = OptionMenu(root,variable, *OptionList)
w.config(font=("Times", 12),background="blue")
#w.pack(side=BOTTOM)
w.place(x=850, y=20)

def callback(*args):
    if variable.get()=="По машинистам":

        scrollbar = Scrollbar(root)
        scrollbar.pack(side=LEFT, fill=Y)
        #scrollbar.place(x=10,y=100)

        listbox = Listbox(root,width=50, yscrollcommand=scrollbar.set)
        l=list_tab_name(PATH)
        #print(l[1][3])
        string=[str(l[0][i])+'-'+str(l[1][i])+'-'+str(l[2][i])+'-'+str(l[3][i]) for i in range(len(l[0]))]
        for i in range(len(string)):
            listbox.insert(END, string[i])
        listbox.pack(side=LEFT,fill=BOTH)
        #listbox.place(x=100,y=100)

        scrollbar.config(command=listbox.yview)

        t2 = Label(root, width=20, text="табельный номер")
        t2.config(font=("Times", 12), background="lightblue")
        # t2.pack(side=LEFT)
        t2.place(x=450, y=550)
        v = StringVar()
        e1 = Entry(root, width=30,textvariable=v)
        e1.place(x=650, y=550)

        def get_value():
            #calc_probability(v.get())
            print(read_data(PATH,'TabNum'))
            if any([read_data(PATH,'TabNum')[i]==int(v.get()) for i in range(len(read_data(PATH,'TabNum')))]):
               print('True')
               calc_probability(v.get())
            #print(v.get())

        Local_button = Button(root, width=20, text="расчет по машинистам",command=get_value)
        #Local_button.bind("<Button-1>", lambda event, f=v.get(): calc_probability(event, f))
        Local_button.config(font=("Times", 12), background="blue")
        #Local_button.pack(side=LEFT)
        Local_button.place(x=850, y=500)
        print(e1.get())

variable.trace("w", callback)


def get_value_none():
    calc_probability('False')
    # print(v.get())

Local_button1 = Button(root,width=20, text="расчет",command=get_value_none)
#Local_button1.bind("<Button-1>", calc_probability)
Local_button1.config(font=("Times", 12),background="blue")
#Local_button1.pack(side=BOTTOM,expand=1)
Local_button1.place(x=850,y=550)
# frame.grid()

root.mainloop()
